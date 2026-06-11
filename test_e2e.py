#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
端到端测试脚本：验证团队周报汇总工具的完整功能
使用方法：python test_e2e.py
"""
import os
import sys
import shutil
import tempfile
import json
import io
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from click.testing import CliRunner
from weekly_report_tool.cli import cli
from weekly_report_tool.models import TeamSummary, ItemType
from weekly_report_tool.storage import StorageManager


def assert_contains(actual, expected_items, label):
    for item in expected_items:
        assert item in actual, f"{label}: 未找到 '{item}'"
    print(f"  ✓ {label}")


def main():
    runner = CliRunner()
    test_dir = tempfile.mkdtemp(prefix="weekly_report_test_")
    os.chdir(test_dir)
    print(f"测试目录: {test_dir}")

    example_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "examples")
    submissions_dir = os.path.join(test_dir, "submissions")
    os.makedirs(submissions_dir)
    for f in os.listdir(example_dir):
        if f.endswith('.txt'):
            shutil.copy(os.path.join(example_dir, f), submissions_dir)

    week = "2026-06-08"
    week_end = "2026-06-14"
    team_name = "测试团队"

    # ====== 1: 帮助信息 ======
    print("\n" + "=" * 60)
    print("测试 1: 帮助信息")
    r = runner.invoke(cli, ['--help'])
    assert r.exit_code == 0
    assert_contains(r.output, ["init", "import", "check", "summary", "export", "history", "rollback"], "顶层 help")
    for cmd in ["init", "import", "check", "summary", "export", "history", "rollback"]:
        r2 = runner.invoke(cli, [cmd, '--help'])
        assert r2.exit_code == 0, f"{cmd} --help 失败"
    print("  ✓ 所有子命令 help 正常")

    # ====== 2: 初始化 ======
    print("\n" + "=" * 60)
    print("测试 2: init")
    r = runner.invoke(cli, ['init', '--team-name', team_name,
                            '--members', '张三,李四,王五,赵六',
                            '--projects', '用户中心,订单系统,商品系统,营销活动,数据分析,基础设施', '-y'])
    assert r.exit_code == 0
    assert_contains(r.output, ["测试团队"], "init")

    # ====== 3: 目录批量导入 ======
    print("\n" + "=" * 60)
    print("测试 3: import 目录导入")
    r = runner.invoke(cli, ['import', '--week', week, '--dir', submissions_dir,
                            '--on-duplicate', 'overwrite', '--no-check'])
    assert r.exit_code == 0
    assert_contains(r.output, ["新增 3"], "目录导入")

    # ====== 4: 重复跳过 ======
    print("\n" + "=" * 60)
    print("测试 4: 重复跳过")
    r = runner.invoke(cli, ['import', '--week', week, '--dir', submissions_dir,
                            '--on-duplicate', 'skip', '--no-check'])
    assert r.exit_code == 0
    assert_contains(r.output, ["跳过 3", "新增 0"], "重复跳过")

    # ====== 5: 重复覆盖 ======
    print("\n" + "=" * 60)
    print("测试 5: 重复覆盖")
    r = runner.invoke(cli, ['import', '--week', week, '--dir', submissions_dir,
                            '--on-duplicate', 'overwrite', '--no-check'])
    assert r.exit_code == 0
    assert_contains(r.output, ["覆盖 3", "新增 0"], "重复覆盖")

    # ====== 6: check ======
    print("\n" + "=" * 60)
    print("测试 6: check")
    r = runner.invoke(cli, ['check', '--week', week])
    assert r.exit_code == 0
    assert_contains(r.output, ["应交：4 人", "已交：3 人", "赵六", "延期", "阻塞", "求助"], "check")

    # ====== 7: summary 按项目 ======
    print("\n" + "=" * 60)
    print("测试 7: summary -g project")
    r = runner.invoke(cli, ['summary', '--week', week, '-g', 'project'])
    assert r.exit_code == 0
    assert_contains(r.output, ["按项目分类", "用户中心", "订单系统"], "按项目")

    # ====== 8: summary 按成员 ======
    print("\n" + "=" * 60)
    print("测试 8: summary -g member")
    r = runner.invoke(cli, ['summary', '--week', week, '-g', 'member'])
    assert r.exit_code == 0
    assert_contains(r.output, ["按成员分类", "张三", "李四", "王五"], "按成员")

    # ====== 9: summary 按负责人 ======
    print("\n" + "=" * 60)
    print("测试 9: summary -g owner")
    r = runner.invoke(cli, ['summary', '--week', week, '-g', 'owner'])
    assert r.exit_code == 0
    assert_contains(r.output, ["按项目负责人分类", "未分配负责人"], "按负责人")

    storage = StorageManager()
    reports = storage.load_all_reports_for_week(week)
    actual_completed = sum(len(rep.get_items_by_type(ItemType.COMPLETED)) for rep in reports.values())
    owner_completed = 0
    for line in r.output.split('\n'):
        if '完成:' in line and '计划:' in line:
            m = re.search(r'完成:(\d+)', line)
            if m:
                owner_completed += int(m.group(1))
    assert owner_completed == actual_completed, f"owner 视图完成数 {owner_completed} != 实际 {actual_completed}"
    print(f"  ✓ owner 视图无重复统计 (完成数={actual_completed})")

    # ====== 10: summary 按风险等级 ======
    print("\n" + "=" * 60)
    print("测试 10: summary -g risk")
    r = runner.invoke(cli, ['summary', '--week', week, '-g', 'risk'])
    assert r.exit_code == 0
    assert_contains(r.output, ["按风险等级分类"], "按风险等级")
    has_risk_cat = any(kw in r.output for kw in ['延期', '阻塞', '求助'])
    assert has_risk_cat, "风险类别未显示"
    print("  ✓ 风险类别正常显示")

    # ====== 11: summary --brief ======
    print("\n" + "=" * 60)
    print("测试 11: summary --brief")
    r = runner.invoke(cli, ['summary', '--week', week, '--brief'])
    assert r.exit_code == 0
    assert_contains(r.output, ["总体统计", "重点关注"], "brief")
    assert "按项目分类" not in r.output
    print("  ✓ 精简模式正常")

    # ====== 12: export 全部格式 ======
    print("\n" + "=" * 60)
    print("测试 12: export --format all")
    r = runner.invoke(cli, ['export', '--week', week, '-f', 'all'])
    assert r.exit_code == 0
    export_dir = os.path.join(test_dir, "weekly_exports")
    files = os.listdir(export_dir)
    week_tag = f"{week}_{week_end}"
    email_files = [f for f in files if "邮件版" in f and week_tag in f]
    group_files = [f for f in files if "群公告版" in f and week_tag in f]
    md_files = [f for f in files if f.endswith(".md") and week_tag in f and "邮件" not in f and "群公告" not in f]
    assert len(email_files) == 1
    assert len(group_files) == 1
    assert len(md_files) == 1
    assert '测试团队' in email_files[0]
    print("  ✓ 三种格式导出正确")

    # ====== 13: 导出内容验证 ======
    print("\n" + "=" * 60)
    print("测试 13: 导出内容完整性")
    with open(os.path.join(export_dir, email_files[0]), encoding="utf-8") as f:
        content = f.read()
    assert_contains(content, ["缺交", "赵六", "需要跟进"], "邮件版")

    with open(os.path.join(export_dir, md_files[0]), encoding="utf-8") as f:
        md_content = f.read()
    assert_contains(md_content, ["# 测试团队周报", "需要跟进"], "Markdown版")

    # ====== 14: 负责人说明 ======
    print("\n" + "=" * 60)
    print("测试 14: 负责人说明追加")
    summary = storage.load_summary(week)
    assert summary is not None
    summary.manual_notes = "本周整体进展顺利，请赵六尽快补交。"
    storage.save_summary(summary)

    r = runner.invoke(cli, ['export', '--week', week, '-f', 'email'])
    assert r.exit_code == 0
    with open(os.path.join(export_dir, email_files[0]), encoding="utf-8") as f:
        content = f.read()
    assert_contains(content, ["【负责人说明】", "本周整体进展顺利"], "负责人说明")

    # ====== 15: 负责人说明持久化 ======
    print("\n" + "=" * 60)
    print("测试 15: 负责人说明持久化（覆盖导入后保留）")
    r = runner.invoke(cli, ['import', '--week', week, '--dir', submissions_dir,
                            '--on-duplicate', 'overwrite', '--no-check'])
    assert r.exit_code == 0
    summary = storage.load_summary(week)
    assert summary is not None, "summary 丢失"
    assert "本周整体进展顺利" in summary.manual_notes, f"manual_notes 丢失: {summary.manual_notes}"
    print("  ✓ 覆盖导入后负责人说明保留")

    # ====== 16: check 后 manual_notes 也保留 ======
    print("\n" + "=" * 60)
    print("测试 16: check 后 manual_notes 保留")
    r = runner.invoke(cli, ['check', '--week', week])
    assert r.exit_code == 0
    summary = storage.load_summary(week)
    assert "本周整体进展顺利" in summary.manual_notes, "check 后 manual_notes 丢失"
    print("  ✓ check 后负责人说明保留")

    # ====== 17: ask 模式 - 跳过 ======
    print("\n" + "=" * 60)
    print("测试 17: ask 模式跳过")
    ask_path = os.path.join(submissions_dir, "ask_test.txt")
    with open(ask_path, "w", encoding="utf-8") as f:
        f.write("姓名：张三\n周期：2026-06-08 ~ 2026-06-14\n\n本周完成：\n- ASK_V2_测试\n")
    r = runner.invoke(cli, ['import', '--week', week, ask_path,
                            '--on-duplicate', 'ask', '--no-check'], input="n\n")
    assert r.exit_code == 0
    report = storage.load_report(week, '张三')
    has_v2 = any('ASK_V2_测试' in it.content for it in report.items)
    assert not has_v2, "跳过后不应有新内容"
    print("  ✓ ask 跳过后旧数据不变")

    # ====== 18: ask 模式 - 覆盖 ======
    print("\n" + "=" * 60)
    print("测试 18: ask 模式覆盖")
    r = runner.invoke(cli, ['import', '--week', week, ask_path,
                            '--on-duplicate', 'ask', '--no-check'], input="y\n")
    assert r.exit_code == 0
    r = runner.invoke(cli, ['summary', '--week', week, '-g', 'member'])
    assert 'ASK_V2_测试' in r.output, "覆盖后 summary 看不到新内容"
    print("  ✓ ask 覆盖后新内容可见")

    # ====== 19: 覆盖后 history 有版本 ======
    print("\n" + "=" * 60)
    print("测试 19: 历史版本")
    r = runner.invoke(cli, ['history', '--week', week])
    assert r.exit_code == 0
    versions_found = '版本:' in r.output or '条目数:' in r.output
    assert versions_found, f"history 无版本记录: {r.output}"
    print("  ✓ history 有历史版本")

    # ====== 20: rollback 回滚 ======
    print("\n" + "=" * 60)
    print("测试 20: rollback 回滚")
    versions = storage.list_versions(week, '张三')
    assert len(versions) > 0, "没有历史版本"
    version_id = versions[0]['version_id']
    r = runner.invoke(cli, ['rollback', '--week', week, '--member', '张三', '--version', version_id])
    assert r.exit_code == 0, f"rollback failed: {r.output}"
    assert '已回滚' in r.output
    print("  ✓ rollback 成功")

    # ====== 21: 回滚后 summary/export 看旧内容 ======
    print("\n" + "=" * 60)
    print("测试 21: 回滚后 export")
    r = runner.invoke(cli, ['export', '--week', week, '-f', 'email'])
    assert r.exit_code == 0
    print("  ✓ 回滚后 export 正常")

    # ====== 22: Excel 多 sheet + 空 sheet ======
    print("\n" + "=" * 60)
    print("测试 22: Excel 多 sheet + 空 sheet")
    from openpyxl import Workbook
    excel_dir = tempfile.mkdtemp(prefix="wr_excel_e2e_")
    old_cwd = os.getcwd()
    os.chdir(excel_dir)
    try:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "张三Excel"
        ws1.append(["姓名：张三Excel"])
        ws1.append(["周期：2026-06-08 ~ 2026-06-14"])
        ws1.append(["本周完成："])
        ws1.append(["- 【用户中心】完成Excel测试"])
        wb.create_sheet("空Sheet")
        ws3 = wb.create_sheet("李四Excel")
        ws3.append(["姓名：李四Excel"])
        ws3.append(["本周完成："])
        ws3.append(["- 【订单系统】完成导出功能"])
        xlsx_path = os.path.join(excel_dir, "test.xlsx")
        wb.save(xlsx_path)

        r = runner.invoke(cli, ['init', '--team-name', 'Excel团队',
                                '--members', '张三Excel,李四Excel,王五',
                                '--projects', '用户中心,订单系统', '-y'])
        assert r.exit_code == 0
        r = runner.invoke(cli, ['import', '--week', '2026-06-08', xlsx_path,
                                '--on-duplicate', 'overwrite', '--no-check'])
        assert r.exit_code == 0
        assert '新增 2' in r.output
        r = runner.invoke(cli, ['check', '--week', '2026-06-08'])
        assert '应交：3 人' in r.output
        assert '已交：2 人' in r.output
        print("  ✓ Excel 多 sheet + 空 sheet 正常")
    finally:
        os.chdir(old_cwd)

    # ====== 23: Excel 表格列式导入 ======
    print("\n" + "=" * 60)
    print("测试 23: Excel 表格列式导入")
    tab_dir = tempfile.mkdtemp(prefix="wr_tab_e2e_")
    os.chdir(tab_dir)
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "周报汇总"
        ws.append(["姓名", "项目", "类型", "内容", "状态"])
        ws.append(["张三Tab", "用户中心", "本周完成", "完成登录模块", ""])
        ws.append(["张三Tab", "用户中心", "风险阻塞", "接口延期，因文档未到位", "延期"])
        ws.append(["李四Tab", "订单系统", "本周完成", "完成退款流程", ""])
        ws.append(["李四Tab", "订单系统", "求助需求", "需要财务确认", ""])
        xlsx_path = os.path.join(tab_dir, "tabular.xlsx")
        wb.save(xlsx_path)

        r = runner.invoke(cli, ['init', '--team-name', 'Tab团队',
                                '--members', '张三Tab,李四Tab,王五Tab',
                                '--projects', '用户中心,订单系统', '-y'])
        assert r.exit_code == 0
        r = runner.invoke(cli, ['import', '--week', '2026-06-08', xlsx_path,
                                '--on-duplicate', 'overwrite', '--no-check'])
        assert r.exit_code == 0
        assert '新增 2' in r.output

        r = runner.invoke(cli, ['check', '--week', '2026-06-08'])
        assert '应交：3 人' in r.output
        assert '已交：2 人' in r.output
        assert '延期' in r.output
        assert '求助' in r.output
        print("  ✓ Excel 表格列式导入正常")
    finally:
        os.chdir(test_dir)

    print("\n" + "=" * 60)
    print("所有 23 项测试通过！")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
