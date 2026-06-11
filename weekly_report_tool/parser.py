import re
from typing import Dict, List, Tuple, Optional
from datetime import datetime, timedelta
from .models import WeeklyReport, ReportItem, ItemType, ItemStatus


SECTION_KEYWORDS = {
    ItemType.COMPLETED: ["本周完成", "已完成", "完成", "本周工作", "done", "completed", "this week"],
    ItemType.PLANNED: ["下周计划", "下周", "计划", "待办", "next week", "todo", "planned"],
    ItemType.RISK: ["风险", "阻塞", "问题", "风险阻塞", "risk", "block", "blocked", "issue"],
    ItemType.HELP: ["求助", "需要", "协助", "帮助", "需求支持", "help", "need", "assistance"]
}

DELAY_KEYWORDS = ["延期", "延迟", "延后", "推迟", "delay", "late", "overdue"]
BLOCK_KEYWORDS = ["阻塞", "卡住", "无法推进", "blocked", "stuck"]


def detect_week_dates(text: str, default_start: Optional[str] = None) -> Tuple[str, str]:
    date_patterns = [
        r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})',
    ]
    dates_found = []
    for pattern in date_patterns:
        matches = re.findall(pattern, text)
        for m in matches:
            try:
                dt = datetime(int(m[0]), int(m[1]), int(m[2]))
                dates_found.append(dt)
            except ValueError:
                continue

    if len(dates_found) >= 2:
        dates_found.sort()
        start = dates_found[0]
        end = dates_found[-1]
    elif len(dates_found) == 1:
        d = dates_found[0]
        start = d - timedelta(days=d.weekday())
        end = start + timedelta(days=6)
    elif default_start:
        try:
            start = datetime.strptime(default_start, "%Y-%m-%d")
            end = start + timedelta(days=6)
        except ValueError:
            today = datetime.now()
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
    else:
        today = datetime.now()
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)

    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def detect_member_name(text: str, filename: str = "") -> str:
    name_patterns = [
        r'(?:姓名|成员|提交人|汇报人|from)[：:]\s*([^\n\r]+)',
        r'(?:报告人|负责人)[:：]\s*([^\n\r]+)',
    ]
    for pattern in name_patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()

    if filename:
        base = os.path.splitext(os.path.basename(filename))[0]
        base = re.sub(r'[\d\-_周周报]', '', base)
        if base:
            return base.strip()

    return "未知成员"


def detect_project_name(line: str, known_projects: List[str] = None) -> str:
    if known_projects:
        for proj in known_projects:
            if proj in line:
                return proj
    proj_match = re.search(r'[【\[]([^\]】]+)[】\]]', line)
    if proj_match:
        return proj_match.group(1)
    return ""


def detect_status(content: str) -> Tuple[ItemStatus, str]:
    status = ItemStatus.NORMAL
    reason = ""
    content_lower = content.lower()
    if re.match(r'^阻塞[:：]?\s*', content) or '：阻塞' in content or '- 阻塞' in content:
        status = ItemStatus.BLOCKED
    else:
        for kw in BLOCK_KEYWORDS:
            if kw in content or kw in content_lower:
                status = ItemStatus.BLOCKED
                break
    if status == ItemStatus.NORMAL:
        if re.match(r'^延期[:：]?\s*', content):
            status = ItemStatus.DELAYED
        else:
            delay_found = False
            for kw in DELAY_KEYWORDS:
                if kw in content or kw in content_lower:
                    delay_found = True
                    break
            if delay_found:
                if not re.search(r'(需要|希望|计划)(进一步)?(优化|改进|调整|完善|测试)', content_lower):
                    status = ItemStatus.DELAYED
        if status == ItemStatus.DELAYED:
            match = re.search(
                r'(?:因|因为|原因|due to|because)[：: ]?([^\n，。,；;]+)',
                content
            )
            if match:
                reason = match.group(1).strip()
    return status, reason


def _split_clean_items(section_text: str) -> List[str]:
    items = []
    lines = section_text.split('\n')
    current = ""
    for line in lines:
        stripped = line.strip()
        if not stripped:
            if current:
                items.append(current.strip())
                current = ""
            continue
        if re.match(r'^[\-\*\d\.\)、]', stripped) or (current and stripped.startswith('  ')):
            if current:
                items.append(current.strip())
            current = re.sub(r'^[\-\*\d\.\)、\s]+', '', stripped)
        else:
            if current:
                current += ' ' + stripped
            else:
                current = stripped
    if current:
        items.append(current.strip())
    return [i for i in items if i]


def parse_text_report(
    text: str,
    filename: str = "",
    known_projects: Optional[List[str]] = None,
    default_week_start: Optional[str] = None
) -> WeeklyReport:
    member_name = detect_member_name(text, filename)
    week_start, week_end = detect_week_dates(text, default_week_start)

    lines = text.split('\n')
    sections: dict = {
        ItemType.COMPLETED: [],
        ItemType.PLANNED: [],
        ItemType.RISK: [],
        ItemType.HELP: []
    }

    current_type: Optional[ItemType] = None
    current_section_lines: List[str] = []
    current_section_project: str = ""

    for line in lines:
        stripped = line.strip()
        matched_type = None
        has_project_tag = bool(re.search(r'[【\[]', stripped))
        ends_with_colon = bool(re.search(r'[：:]\s*$', stripped))

        header_project = detect_project_name(stripped, known_projects) if has_project_tag else ""

        header_without_project = re.sub(r'[【\[][^\]】]+[】\]]', '', stripped).strip()
        header_without_project = re.sub(r'^[一二三四五六七八九十\d]+[、\.\)：:\s]+', '', header_without_project)
        header_without_project = re.sub(r'^[#\*\-\s]+', '', header_without_project)

        is_short_header = len(header_without_project) <= 15 or ends_with_colon
        if has_project_tag:
            is_short_header = len(header_without_project) <= 6 and bool(header_without_project)

        for itype, keywords in SECTION_KEYWORDS.items():
            for kw in keywords:
                kw_lower = kw.lower()
                stripped_lower = stripped.lower()
                kw_found = kw in stripped or kw_lower in stripped_lower
                if not kw_found:
                    continue
                if ends_with_colon:
                    matched_type = itype
                    break
                if is_short_header:
                    hdr = header_without_project if has_project_tag else stripped
                    hdr_clean = re.sub(r'^[一二三四五六七八九十\d]+[、\.\)：:\s]+', '', hdr)
                    hdr_clean = re.sub(r'^[#\*\-\s]+', '', hdr_clean)
                    if (hdr_clean == kw or hdr_clean.lower() == kw_lower or
                            kw in hdr_clean or kw_lower in hdr_clean.lower()):
                        matched_type = itype
                        break
            if matched_type:
                break

        if matched_type:
            if current_type is not None:
                section_text = '\n'.join(current_section_lines)
                items_text = _split_clean_items(section_text)
                sections[current_type].extend([(c, current_section_project) for c in items_text])
            current_type = matched_type
            current_section_lines = []
            current_section_project = header_project
            colon_idx = max(stripped.find('：'), stripped.find(':'))
            if colon_idx > 0:
                rest = stripped[colon_idx + 1:].strip()
                if rest:
                    current_section_lines.append(rest)
        elif current_type is not None:
            current_section_lines.append(line)

    if current_type is not None:
        section_text = '\n'.join(current_section_lines)
        items_text = _split_clean_items(section_text)
        sections[current_type].extend([(c, current_section_project) for c in items_text])

    EMPTY_CONTENT_PATTERNS = [
        r'^无$', r'^暂无$', r'^没有$', r'^none$', r'^n/a$', r'^na$',
        r'^无。$', r'^暂无。$', r'^没有。$',
    ]

    report_items: List[ReportItem] = []
    for itype, item_contents in sections.items():
        for content, section_project in item_contents:
            content_stripped = content.strip()
            is_empty = False
            for pat in EMPTY_CONTENT_PATTERNS:
                if re.match(pat, content_stripped, re.IGNORECASE):
                    is_empty = True
                    break
            if is_empty:
                continue

            project = detect_project_name(content, known_projects)
            if not project and section_project:
                project = section_project
            status, delay_reason = detect_status(content)
            deadline_match = re.search(
                r'(?:截止|deadline|ddl)[：: ]?(\d{4}[-/\.]?\d{0,2}[-/\.]?\d{0,2})',
                content,
                re.IGNORECASE
            )
            deadline = deadline_match.group(1) if deadline_match else ""
            assignee_match = re.search(
                r'(?:负责人|对接人|assigned to|@)[：: ]?([^\n，。,；\s]+)',
                content,
                re.IGNORECASE
            )
            assignee = assignee_match.group(1) if assignee_match else member_name
            report_items.append(ReportItem(
                item_type=itype,
                content=content,
                project=project,
                status=status,
                delay_reason=delay_reason,
                deadline=deadline,
                assignee=assignee
            ))

    return WeeklyReport(
        member_name=member_name,
        week_start=week_start,
        week_end=week_end,
        items=report_items
    )


TYPE_COLUMN_KEYWORDS = {
    ItemType.COMPLETED: ["完成", "已完成", "本周完成", "本周工作", "done", "completed"],
    ItemType.PLANNED: ["计划", "下周计划", "待办", "planned", "todo", "next"],
    ItemType.RISK: ["风险", "阻塞", "风险阻塞", "问题", "risk", "block", "issue"],
    ItemType.HELP: ["求助", "协助", "帮助", "help", "need", "assistance"],
}

HEADER_ALIASES = {
    "name": ["姓名", "成员", "提交人", "汇报人", "name", "member"],
    "week": ["周期", "周", "日期", "week", "date", "period"],
    "project": ["项目", "项目名", "project"],
    "type": ["类型", "分类", "type", "category"],
    "content": ["内容", "描述", "事项", "content", "description", "detail"],
    "status": ["状态", "status"],
}


def _detect_tabular_headers(row_values: list) -> Optional[dict]:
    header_map = {}
    for col_idx, val in enumerate(row_values):
        if val is None:
            continue
        val_str = str(val).strip().lower()
        if not val_str:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if field in header_map:
                continue
            for alias in aliases:
                if alias == val_str or alias in val_str:
                    header_map[field] = col_idx
                    break
    if "content" in header_map and len(header_map) >= 2:
        return header_map
    if "name" in header_map and "type" in header_map:
        return header_map
    if "name" in header_map and "content" in header_map:
        return header_map
    return None


def _parse_type_from_value(val: str) -> Optional[ItemType]:
    val_lower = val.strip().lower()
    for itype, keywords in TYPE_COLUMN_KEYWORDS.items():
        for kw in keywords:
            if kw == val_lower or kw in val_lower:
                return itype
    return None


def parse_excel_tabular(sheet, known_projects: Optional[List[str]] = None,
                        default_week_start: Optional[str] = None) -> List[WeeklyReport]:
    member_reports: Dict[str, WeeklyReport] = {}
    header_map = None
    default_week_end = None
    if default_week_start:
        try:
            from datetime import datetime, timedelta
            ws = datetime.strptime(default_week_start, "%Y-%m-%d")
            default_week_end = (ws + timedelta(days=6)).strftime("%Y-%m-%d")
        except ValueError:
            pass

    for row in sheet.iter_rows(values_only=True):
        row_strs = [str(c).strip() if c is not None else "" for c in row]
        non_empty = [s for s in row_strs if s]
        if not non_empty:
            continue

        if header_map is None:
            candidate = _detect_tabular_headers(row)
            if candidate:
                header_map = candidate
                continue
            else:
                continue

        content_val = row_strs[header_map.get("content", -1)] if "content" in header_map else ""
        if not content_val.strip():
            continue

        name_val = row_strs[header_map["name"]].strip() if "name" in header_map else ""
        if not name_val:
            continue

        project_val = row_strs[header_map.get("project", -1)].strip() if "project" in header_map else ""
        if not project_val and known_projects:
            for kp in known_projects:
                if kp in content_val:
                    project_val = kp
                    break

        type_val = row_strs[header_map.get("type", -1)].strip() if "type" in header_map else ""
        item_type = _parse_type_from_value(type_val) if type_val else ItemType.COMPLETED

        if item_type is None:
            item_type = ItemType.COMPLETED

        week_val = row_strs[header_map.get("week", -1)].strip() if "week" in header_map else ""
        week_start = default_week_start or ""
        week_end = default_week_end or ""
        if week_val:
            ws, we = detect_week_dates(week_val, default_week_start)
            week_start = ws
            week_end = we

        status_val = row_strs[header_map.get("status", -1)].strip() if "status" in header_map else ""
        if status_val:
            status_lower = status_val.lower()
            if any(k in status_lower for k in ["阻塞", "blocked", "stuck"]):
                item_status = ItemStatus.BLOCKED
                delay_reason = ""
            elif any(k in status_lower for k in ["延期", "延迟", "delayed", "late"]):
                item_status = ItemStatus.DELAYED
                delay_reason = ""
            else:
                item_status, delay_reason = detect_status(content_val)
        else:
            item_status, delay_reason = detect_status(content_val)

        if name_val not in member_reports:
            member_reports[name_val] = WeeklyReport(
                member_name=name_val,
                week_start=week_start,
                week_end=week_end,
                items=[]
            )

        member_reports[name_val].items.append(ReportItem(
            item_type=item_type,
            content=content_val,
            project=project_val,
            status=item_status,
            delay_reason=delay_reason,
            assignee=name_val
        ))

    return list(member_reports.values())


def parse_excel_report(
    filepath: str,
    known_projects: Optional[List[str]] = None,
    default_week_start: Optional[str] = None
) -> List[WeeklyReport]:
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    reports: List[WeeklyReport] = []
    default_name_hints = {"sheet", "sheet1", "sheet2", "sheet3", "工作表", "数据"}

    for sheet in wb.worksheets:
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        header_map = _detect_tabular_headers(list(first_row)) if first_row else None

        if header_map:
            sheet_reports = parse_excel_tabular(
                sheet, known_projects=known_projects,
                default_week_start=default_week_start
            )
            reports.extend(sheet_reports)
            continue

        text_parts = []
        for row in sheet.iter_rows(values_only=True):
            row_strs = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if row_strs:
                text_parts.append(" | ".join(row_strs))

        text = "\n".join(text_parts)
        if not text.strip():
            continue

        report = parse_text_report(
            text,
            filename=os.path.basename(filepath),
            known_projects=known_projects,
            default_week_start=default_week_start
        )

        sheet_title = sheet.title.strip() if sheet.title else ""
        if report.member_name == "未知成员" and sheet_title and sheet_title.lower() not in default_name_hints:
            report.member_name = sheet_title

        if len(report.items) == 0 and report.member_name == "未知成员":
            continue

        reports.append(report)

    return reports


import os
